import openpyxl
from tabulate import tabulate

inventario = openpyxl.load_workbook("Inventario.xlsx")
inventarioRows = 16
reportesRows = 8

shenames = inventario.sheetnames

worksheet1 = inventario[shenames[0]]
worksheet2 = inventario[shenames[1]]

def consultarInventario():
    tablaInventario = []
    for i in range(inventarioRows):
        filaInventario = []
        for cell in list(worksheet1.rows)[i]:
            if cell.value != None:
                filaInventario.append(cell.value)
        tablaInventario.append(filaInventario)
    print(tabulate(tablaInventario))

def llegadaProductos():
    arregloCodigos = []
    for cell in list(worksheet1.columns)[0]:
        arregloCodigos.append(cell.value)
    while True:
        print("\n--------------\n1. Ingresar producto\n2. Regresar")
        decision = int(input("Ingrese la acción que desea realizar: "))
        if decision == 1:
            consultarInventario()
            ingresarCodigo = int(input("Ingrese el codigo del producto: "))
            if ingresarCodigo in arregloCodigos:
                ingresarCantidad = int(input("Ingrese la cantidad de productos que llegaron: "))
                fila = arregloCodigos.index(ingresarCodigo)
                celda = worksheet1.cell(fila + 1,4).value
                productoAgregado = celda + ingresarCantidad
                worksheet1.cell(fila + 1, 4, productoAgregado)
        elif decision == 2:
            break

def mostrarReportes():
    tablaReportes = []
    for i in range(reportesRows):
        filaReporte = []
        for cell in list(worksheet2.rows)[i]:
            filaReporte.append(cell.value)
        tablaReportes.append(filaReporte)
    print(tabulate(tablaReportes))

def reportesEspecificos():
    arregloCodigos = []
    for cell in list(worksheet1.columns)[0]:
        arregloCodigos.append(cell.value)
    arregloMatriculas = []
    for cell in list(worksheet2.columns)[0]:
        arregloMatriculas.append(cell.value)
    while True:
        print("\n--------------\n1. Reporte por vendedor\n2. Reporte por producto\n3. Regresar")
        decision = int(input("Ingrese la acción que desea realizar: "))
        if decision == 1:
            ingresarMatricula = int(input("Ingresa la matricula del vendedor: "))
            tablaReporte = []
            arregloTop = []
            for cell in list(worksheet2.rows)[0]:
                arregloTop.append(cell.value)
            fila = arregloMatriculas.index(ingresarMatricula)
            arregloVendedor = []
            for cell in list(worksheet2.rows)[fila]:
                arregloVendedor.append(cell.value)
            tablaReporte.append(arregloTop)
            tablaReporte.append(arregloVendedor)
            print(tabulate(tablaReporte))
        if decision == 2:
            ingresarCodigo = int(input("Ingresa el código del producto: "))
            tablaReporte = []
            arregloLeft1 = []
            arregloLeft2 = []
            for cell in list(worksheet2.columns)[0]:
                arregloLeft1.append(cell.value)
            for cell in list(worksheet2.columns)[1]:
                arregloLeft2.append(cell.value)
            fila = arregloCodigos.index(ingresarCodigo) + 2
            arregloProducto = []
            for cell in list(worksheet2.columns)[fila]:
                arregloProducto.append(cell.value)
            tablaReporte.append(arregloLeft1)
            tablaReporte.append(arregloLeft2)
            tablaReporte.append(arregloProducto)
            print(tabulate(tablaReporte))
        if decision == 3:
            break

def administrador():
    while True:
        print("\nBienvenido al menu de administrador\n--------------\n1. Registrar llegada de productos\n2. Consultar el inventario\n3. Mostrar reporte general de ventas\n4. Mostrar reportes individuales de ventas\n5. Regresar")
        decision = int(input("Ingrese la accion que desea realizar: "))
        if decision == 1:
            llegadaProductos()
        elif decision == 2:
            consultarInventario()
        elif decision == 3:
            mostrarReportes()
        elif decision == 4:
            reportesEspecificos()
        elif decision == 5:
            break

def registrarVenta():
    arregloCodigos = []
    for cell in list(worksheet1.columns)[0]:
        arregloCodigos.append(cell.value)
    arregloMatriculas = []
    for cell in list(worksheet2.columns)[0]:
        arregloMatriculas.append(cell.value)
    while True:
        print("\n--------------\n1. Ingresar producto\n2. Regresar")
        decision = int(input("Ingrese la acción que desea realizar: "))
        if decision == 1:
            ingresarCodigo = int(input("Ingrese el codigo del producto: "))
            ingresarMatricula = int(input("Ingrese su número de empleado: "))
            if ingresarCodigo in arregloCodigos:
                if ingresarMatricula in arregloMatriculas:
                    ingresarCantidad = int(input("Ingrese la cantidad de productos que vendiste: "))
                    fila = arregloMatriculas.index(ingresarMatricula) + 1
                    columna = arregloCodigos.index(ingresarCodigo) + 2
                    celdaProducto = worksheet2.cell(fila,columna).value
                    productoVendido = celdaProducto + ingresarCantidad
                    worksheet2.cell(fila,columna,productoVendido)
                    celdaVendido = worksheet2.cell(8,columna).value
                    operacionVendido = celdaVendido + ingresarCantidad
                    worksheet2.cell(8,columna,operacionVendido)
                    filaInventario = arregloCodigos.index(ingresarCodigo)
                    celdaInventario = worksheet1.cell(filaInventario + 1,4).value
                    print(celdaInventario)
                    productoRestado = celdaInventario - ingresarCantidad
                    print(productoRestado)
                    worksheet1.cell(filaInventario + 1, 4, productoRestado)   
        elif decision == 2:
            break


def vendedor():
     while(True):
        print("\nBienvenido al menu de vendedor\n--------------\n1. Registrar una venta\n2. Regresar")
        decision = int(input("Ingrese la accion que desea realizar: "))
        if decision == 1:
            registrarVenta()
        elif decision == 2:
            break

def main():
    while(True):
        print("\nPROGRAMA INVENTARIO\n--------------\n1. Administrador\n2. Vendedor\n3. Salir")
        sesion = int(input("Ingrese como desea iniciar: "))

        if sesion == 1:
            administrador()
        elif sesion == 2:
            vendedor()
        elif sesion == 3:
            print("Adios!")
            break

main()